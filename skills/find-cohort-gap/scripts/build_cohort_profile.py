#!/usr/bin/env python3
"""Build a cohort profile from a local codebook (and optional domain context).

`/find-cohort-gap` used to start from a *named* database — NHIS, UK Biobank, and the
handful of registries the skill knows about. Most researchers do not have one of those.
They have an institutional registry, a single-centre EMR export, or a specialty cohort,
described by a data dictionary nobody else has ever seen (issue #69).

This is the input layer that lets them in. It reads their codebook and emits the same
cohort profile the skill already consumes, so everything downstream — the intersection
matrix, saturation scan, 6-pattern scoring — is unchanged.

WHY A SCRIPT AND NOT JUST READING THE FILE. A language model asked to "summarise this
codebook" will paraphrase a variable name, merge two variables that look alike, or
quietly invent one that the cohort does not have — and every downstream claim inherits
it. So the variable inventory is *enumerated*, never generated: each variable is copied
verbatim from the file and carries its provenance (`file:row`), which is exactly the
dictionary-first discipline that a reviewer expects of a DB-backed study.

WHAT IT REFUSES TO DO. A codebook lists variables; it does not state the sample size,
the enrollment window, the follow-up duration, or the IRB status. Those are emitted as
`[UNKNOWN - ask the user]`, never guessed — a fabricated N is worse than a missing one,
because it survives all the way to a feasibility gate that then passes for the wrong
reason.

WHAT IT INFERS (and shows its work). Two structural facts *are* derivable from variable
names alone, and both feed patterns the skill already scores:

  * serial / repeated measures (`bp_v1`, `bp_v2`, `visit2_hba1c`, ...) -> P1 Longitudinal
    Advantage. Reported as the actual variable groups, so the claim is auditable.
  * endpoint-like variables (`death_date`, `cvd_event`, `cancer_incidence`) -> P2
    Endpoint Upgrade.

Every cluster assignment records the keyword that triggered it, and anything that matches
nothing is left `unclassified` rather than forced into a bucket.

Usage:
    build_cohort_profile.py --codebook dict.csv [--codebook more.xlsx ...] \\
        [--context review.pdf --context https://example.org/guideline] \\
        [--cohort-name "Institutional CT registry"] [--out-dir .]

Formats: .csv / .tsv / .json / .md / .txt (stdlib), .xlsx (needs openpyxl),
.pdf (needs `pdftotext` from poppler). Context URLs are fetched with stdlib urllib;
a paywalled or JavaScript-rendered page will not extract, and that is reported rather
than papered over.
"""

from __future__ import annotations

import argparse
import csv
import html.parser
import json
import re
import shutil
import subprocess
import sys
import urllib.error
import urllib.request
from pathlib import Path

UNKNOWN = "[UNKNOWN - ask the user]"

# Variable-name lexicons. Each cluster maps to the template's category names, and the
# matched keyword is recorded so a wrong assignment is visible instead of silent.
#
# Short keys are matched as WHOLE TOKENS, not substrings (see `classify`). Substring
# matching on a two-letter abbreviation is a false-positive machine: `us` (ultrasound)
# fires on `statin_use`, and `age` fires on `storage_temp`. Long keys keep substring
# matching so that `smok` still catches `smoking_status`.
CLUSTERS: dict[str, tuple[str, ...]] = {
    "demographics": (
        "age", "sex", "gender", "birth", "race", "ethnic", "income", "educat",
        "marital", "occupation", "smok", "alcohol", "drink", "exercise", "activity",
    ),
    "anthropometry": ("height", "weight", "bmi", "waist", "hip", "circumf", "body_fat", "muscle"),
    "vital_signs": ("sbp", "dbp", "bp", "blood_pressure", "pulse", "heart_rate", "hr", "resp_rate", "temp"),
    "laboratory": (
        "glucose", "glu", "hba1c", "chol", "lipid", "ldl", "hdl", "triglyc", "tg", "ast", "alt",
        "ggt", "bilirubin", "albumin", "creatinin", "egfr", "bun", "uric", "crp", "esr",
        "hb", "hgb", "hct", "wbc", "rbc", "platelet", "psa", "cea", "afp", "ca19", "tsh",
        "insulin", "lab",
    ),
    "imaging": (
        "ct", "mri", "xray", "x_ray", "cxr", "ultrasound", "us", "sono", "dexa",
        "dxa", "mammo", "echo", "angio", "pet", "cac", "calcium_score", "imaging", "radiol",
    ),
    "questionnaire": (
        "phq", "gad", "psqi", "ipaq", "sf36", "sf_36", "eq5d", "eq_5d", "qol", "questionn",
        "survey", "scale", "sleep", "diet", "food_freq", "ffq", "stress", "depress",
    ),
    "medication": ("med", "drug", "rx", "prescri", "statin", "antihyp", "therapy", "treat"),
    "procedure": ("surg", "operat", "procedur", "biopsy", "resect", "ablat", "stent", "pci", "cabg", "intervention"),
    "diagnosis": ("icd", "diag", "dx", "disease", "hypertension", "diabetes", "history_of", "hx", "comorbid"),
    "identifier_admin": ("id", "key", "index", "seq", "code", "visit_date", "site", "center"),
}

# Below this length a keyword is matched as a whole token, never as a substring.
TOKEN_ONLY = 4

# Endpoint-like variables. A cohort's value proposition is usually its hard endpoints, and
# the skill scores that (P2), so surface them explicitly.
ENDPOINT_HINTS = (
    "death", "mortal", "expire", "died", "survival", "cancer", "malign", "incid",
    "cvd", "chd", "mace", "stroke", "myocard", "mi", "infarct", "event", "outcome",
    "hospital", "admission", "readmit", "recurrence", "progression", "relapse",
)

# Serial / repeated measures: <stem><separator><index>, or an explicit visit/wave marker.
SERIAL_SUFFIX = re.compile(r"^(?P<stem>.+?)[ _\-.]?(?:v|t|w|wave|visit|yr|y|round|r|time)?(?P<idx>\d{1,2})$", re.I)
SERIAL_PREFIX = re.compile(r"^(?:v|t|w|wave|visit|yr|y|round|time)(?P<idx>\d{1,2})[ _\-.](?P<stem>.+)$", re.I)

# A column in a codebook that holds the variable NAME (as opposed to its description).
NAME_COL = re.compile(r"^\s*(variable|var|var_?name|name|field|field_?name|column|col|item|code)\s*$", re.I)
DESC_COL = re.compile(r"^\s*(desc|description|label|definition|meaning|explanation|comment|note|한글|설명)", re.I)


class _Strip(html.parser.HTMLParser):
    """Minimal HTML -> text. Not a browser: a JS-rendered page yields nothing, and the
    caller is told so rather than being handed an empty 'context'."""

    def __init__(self) -> None:
        super().__init__()
        self.chunks: list[str] = []
        self._skip = 0

    def handle_starttag(self, tag, attrs):
        if tag in ("script", "style", "nav", "footer"):
            self._skip += 1

    def handle_endtag(self, tag):
        if tag in ("script", "style", "nav", "footer") and self._skip:
            self._skip -= 1

    def handle_data(self, data):
        if not self._skip and data.strip():
            self.chunks.append(data.strip())


# --------------------------------------------------------------------------------------
# Variable extraction — one function per format, all returning (name, description, where)
# --------------------------------------------------------------------------------------

Var = tuple[str, str, str]


def _from_delimited(path: Path, delim: str) -> list[Var]:
    """A .csv/.tsv is either a CODEBOOK (rows are variables) or a DATA export (columns are
    variables). Decide by whether a header column names a variable column."""
    with path.open(newline="", encoding="utf-8-sig", errors="replace") as fh:
        rows = list(csv.reader(fh, delimiter=delim))
    if not rows:
        return []
    header = rows[0]
    name_idx = next((i for i, h in enumerate(header) if NAME_COL.match(h or "")), None)

    if name_idx is None:  # data export: the header row IS the variable list
        return [
            (h.strip(), "", f"{path.name}:1 (column {i + 1})")
            for i, h in enumerate(header)
            if h and h.strip()
        ]

    desc_idx = next((i for i, h in enumerate(header) if DESC_COL.match(h or "")), None)
    out: list[Var] = []
    for r, row in enumerate(rows[1:], start=2):
        if name_idx >= len(row) or not (row[name_idx] or "").strip():
            continue
        desc = row[desc_idx].strip() if desc_idx is not None and desc_idx < len(row) else ""
        out.append((row[name_idx].strip(), desc, f"{path.name}:{r}"))
    return out


def _from_json(path: Path) -> list[Var]:
    data = json.loads(path.read_text(encoding="utf-8"))
    out: list[Var] = []
    if isinstance(data, dict):
        for i, (k, v) in enumerate(data.items(), start=1):
            desc = v if isinstance(v, str) else (v.get("description", "") if isinstance(v, dict) else "")
            out.append((str(k), str(desc), f"{path.name}:key {i}"))
    elif isinstance(data, list):
        for i, item in enumerate(data, start=1):
            if isinstance(item, str):
                out.append((item, "", f"{path.name}:item {i}"))
            elif isinstance(item, dict):
                name = next((item[k] for k in ("variable", "var", "name", "field", "column") if k in item), None)
                if name:
                    desc = next((item[k] for k in ("description", "desc", "label", "definition") if k in item), "")
                    out.append((str(name), str(desc), f"{path.name}:item {i}"))
    return out


def _from_markdown(text: str, origin: str) -> list[Var]:
    """A markdown/plain codebook: a pipe table, or `var` — description lines."""
    out: list[Var] = []
    for i, line in enumerate(text.splitlines(), start=1):
        s = line.strip()
        if not s or set(s) <= set("|-: "):  # separator row
            continue
        if s.startswith("|"):
            cells = [c.strip().strip("`*") for c in s.strip("|").split("|")]
            if len(cells) >= 1 and cells[0] and not NAME_COL.match(cells[0]):
                out.append((cells[0], cells[1] if len(cells) > 1 else "", f"{origin}:{i}"))
            continue
        m = re.match(r"^[-*+]?\s*`([^`]+)`\s*[-–—:]?\s*(.*)$", s)
        if m:
            out.append((m.group(1).strip(), m.group(2).strip(), f"{origin}:{i}"))
    return out


def _from_xlsx(path: Path) -> list[Var]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError:
        raise SystemExit(
            f"reading {path.name} needs openpyxl (`pip install openpyxl`). "
            "Or export the sheet to .csv and pass that — the CSV path is stdlib-only."
        )
    ws = load_workbook(path, read_only=True, data_only=True).active
    rows = [[("" if c is None else str(c)) for c in row] for row in ws.iter_rows(values_only=True)]
    if not rows:
        return []
    header = rows[0]
    name_idx = next((i for i, h in enumerate(header) if NAME_COL.match(h or "")), None)
    if name_idx is None:
        return [(h.strip(), "", f"{path.name}:1 (column {i + 1})") for i, h in enumerate(header) if h.strip()]
    desc_idx = next((i for i, h in enumerate(header) if DESC_COL.match(h or "")), None)
    out: list[Var] = []
    for r, row in enumerate(rows[1:], start=2):
        if name_idx >= len(row) or not row[name_idx].strip():
            continue
        desc = row[desc_idx].strip() if desc_idx is not None and desc_idx < len(row) else ""
        out.append((row[name_idx].strip(), desc, f"{path.name}:{r}"))
    return out


def _pdf_text(path: Path) -> str:
    if not shutil.which("pdftotext"):
        raise SystemExit(
            f"reading {path.name} needs `pdftotext` (poppler: `brew install poppler` / "
            "`apt install poppler-utils`). Or convert the PDF to .md / .txt and pass that."
        )
    p = subprocess.run(["pdftotext", "-layout", str(path), "-"], capture_output=True, text=True)
    if p.returncode != 0:
        raise SystemExit(f"pdftotext failed on {path.name}: {p.stderr.strip()}")
    return p.stdout


def read_codebook(path: Path) -> list[Var]:
    if not path.is_file():
        raise SystemExit(f"not found: {path}")
    suf = path.suffix.lower()
    if suf == ".csv":
        return _from_delimited(path, ",")
    if suf in (".tsv", ".tab"):
        return _from_delimited(path, "\t")
    if suf == ".json":
        return _from_json(path)
    if suf == ".xlsx":
        return _from_xlsx(path)
    if suf == ".pdf":
        return _from_markdown(_pdf_text(path), path.name)
    if suf in (".md", ".markdown", ".txt"):
        return _from_markdown(path.read_text(encoding="utf-8", errors="replace"), path.name)
    raise SystemExit(f"unsupported codebook format: {path.suffix} ({path.name})")


# --------------------------------------------------------------------------------------
# Structure inference — clusters, serial groups, endpoints
# --------------------------------------------------------------------------------------


def _tokens(s: str) -> set[str]:
    return set(re.split(r"[^a-z0-9]+", s.lower())) - {""}


def _hits(key: str, hay: str, toks: set[str]) -> bool:
    """A short key must BE a token; a long key may appear anywhere.

    This is the whole defence against the abbreviation false positive: `us` matching
    `statin_use`, `age` matching `storage_temp`, `id` matching `lipid`.
    """
    if len(key) >= TOKEN_ONLY or "_" in key:
        return key in hay
    return key in toks


def classify(name: str, desc: str) -> tuple[str, str]:
    """Return (cluster, matched_keyword). Unmatched stays 'unclassified' — a variable is
    never forced into a bucket to make the map look complete."""
    hay = f"{name} {desc}".lower()
    toks = _tokens(hay)
    for cluster, keys in CLUSTERS.items():
        for k in keys:
            if _hits(k, hay, toks):
                return cluster, k
    return "unclassified", ""


def is_endpoint(name: str, desc: str) -> str:
    hay = f"{name} {desc}".lower()
    toks = _tokens(hay)
    return next((k for k in ENDPOINT_HINTS if _hits(k, hay, toks)), "")


def serial_groups(names: list[str]) -> dict[str, list[str]]:
    """Group variables that look like the same measurement repeated over time.

    Only a stem seen with >= 2 distinct indices counts: a lone `visit1_bp` is not evidence
    of serial data, and claiming otherwise would hand P1 a free point it has not earned.
    """
    stems: dict[str, dict[str, str]] = {}
    for n in names:
        for rx in (SERIAL_PREFIX, SERIAL_SUFFIX):
            m = rx.match(n)
            if m:
                stem = m.group("stem").strip("_- .").lower()
                if stem:
                    stems.setdefault(stem, {})[m.group("idx")] = n
                break
    return {
        stem: [v for _, v in sorted(idx.items(), key=lambda kv: int(kv[0]))]
        for stem, idx in stems.items()
        if len(idx) >= 2
    }


# --------------------------------------------------------------------------------------
# Domain context (a review, a guideline page) — extracted, never summarised here
# --------------------------------------------------------------------------------------


def read_context(src: str) -> tuple[str, str]:
    """Return (text, provenance). The LLM reads this; the script only fetches it."""
    if src.startswith(("http://", "https://")):
        try:
            req = urllib.request.Request(src, headers={"User-Agent": "medsci-skills/find-cohort-gap"})
            with urllib.request.urlopen(req, timeout=30) as r:  # noqa: S310 - user-supplied context URL
                raw = r.read().decode("utf-8", errors="replace")
        except (urllib.error.URLError, TimeoutError, ValueError) as exc:
            raise SystemExit(
                f"could not fetch {src}: {exc}\n"
                "If the page is paywalled or JavaScript-rendered (UpToDate and many guideline "
                "portals are), save it as PDF/markdown and pass the file instead."
            )
        p = _Strip()
        p.feed(raw)
        text = "\n".join(p.chunks)
        if len(text.split()) < 50:
            raise SystemExit(
                f"{src} yielded almost no text ({len(text.split())} words) — it is probably "
                "JavaScript-rendered or paywalled. Save it as PDF/markdown and pass the file."
            )
        return text, src
    path = Path(src)
    if not path.is_file():
        raise SystemExit(f"not found: {src}")
    if path.suffix.lower() == ".pdf":
        return _pdf_text(path), path.name
    return path.read_text(encoding="utf-8", errors="replace"), path.name


# --------------------------------------------------------------------------------------


def build(codebooks: list[Path], contexts: list[str], cohort_name: str | None) -> tuple[dict, str]:
    variables: list[dict] = []
    seen: set[str] = set()
    for cb in codebooks:
        for name, desc, where in read_codebook(cb):
            if name.lower() in seen:
                continue
            seen.add(name.lower())
            cluster, kw = classify(name, desc)
            variables.append(
                {
                    "name": name,
                    "description": desc,
                    "source": where,
                    "cluster": cluster,
                    "matched_keyword": kw,
                    "endpoint_hint": is_endpoint(name, desc),
                }
            )

    names = [v["name"] for v in variables]
    serial = serial_groups(names)
    endpoints = [v for v in variables if v["endpoint_hint"]]

    by_cluster: dict[str, list[str]] = {}
    for v in variables:
        by_cluster.setdefault(v["cluster"], []).append(v["name"])

    ctx: list[dict] = []
    for c in contexts:
        text, prov = read_context(c)
        ctx.append({"source": prov, "words": len(text.split()), "text": text})

    profile = {
        "cohort_name": cohort_name or UNKNOWN,
        "codebooks": [str(c) for c in codebooks],
        "n_variables": len(variables),
        "variables": variables,
        "clusters": {k: sorted(v) for k, v in sorted(by_cluster.items())},
        "serial_groups": serial,
        "endpoint_candidates": [v["name"] for v in endpoints],
        "context_documents": [{"source": c["source"], "words": c["words"]} for c in ctx],
        # Not derivable from a codebook. Guessing any of these corrupts the feasibility gate.
        "must_ask_user": {
            "n_baseline": UNKNOWN,
            "n_with_followup": UNKNOWN,
            "enrollment_period": UNKNOWN,
            "followup_duration": UNKNOWN,
            "measurement_intervals": UNKNOWN,
            "irb_status": UNKNOWN,
            "existing_publications": UNKNOWN,
            "known_limitations": UNKNOWN,
        },
    }
    return profile, "\n\n".join(f"# Context: {c['source']}\n\n{c['text']}" for c in ctx)


def render_markdown(p: dict) -> str:
    L = [
        "# Cohort Profile (auto-generated)",
        "",
        f"**Cohort name:** {p['cohort_name']}",
        f"**Codebook(s):** {', '.join(p['codebooks']) or UNKNOWN}",
        f"**Variables enumerated:** {p['n_variables']}",
        "",
        "> Every variable below is copied verbatim from the codebook and carries its source",
        "> location. Nothing here is inferred except the cluster assignment and the serial /",
        "> endpoint flags, each of which shows the keyword that triggered it.",
        "",
        "## Variable Cluster Map",
        "",
        "| Cluster | N | Variables |",
        "|---------|--:|-----------|",
    ]
    for cluster, vs in p["clusters"].items():
        shown = ", ".join(f"`{v}`" for v in vs[:12])
        if len(vs) > 12:
            shown += f", … (+{len(vs) - 12})"
        L.append(f"| {cluster} | {len(vs)} | {shown} |")

    L += ["", "## Serial / repeated measures (evidence for P1 Longitudinal Advantage)", ""]
    if p["serial_groups"]:
        L += ["| Measurement | Timepoints | Variables |", "|-------------|-----------:|-----------|"]
        for stem, vs in sorted(p["serial_groups"].items()):
            L.append(f"| {stem} | {len(vs)} | {', '.join(f'`{v}`' for v in vs)} |")
    else:
        L.append("None detected from variable names. If the cohort *does* have repeated measures,")
        L.append("say so — the naming convention may simply not encode the timepoint.")

    L += ["", "## Endpoint candidates (evidence for P2 Endpoint Upgrade)", ""]
    if p["endpoint_candidates"]:
        L += [", ".join(f"`{v}`" for v in p["endpoint_candidates"])]
    else:
        L.append("None detected. Without a hard endpoint, P2 cannot score and the cohort is")
        L.append("limited to cross-sectional questions — confirm before proceeding.")

    if p["context_documents"]:
        L += ["", "## Domain context supplied", ""]
        for c in p["context_documents"]:
            L.append(f"- {c['source']} ({c['words']:,} words) — see `context_extract.md`")

    L += [
        "",
        "## Not derivable from a codebook — ASK THE USER before Phase 2",
        "",
        "A data dictionary lists variables. It does not state any of the following, and a",
        "guessed value here would silently pass the Phase 5 feasibility gate for the wrong reason.",
        "",
    ]
    L += [f"- **{k.replace('_', ' ')}:** {v}" for k, v in p["must_ask_user"].items()]
    return "\n".join(L) + "\n"


def main() -> int:
    ap = argparse.ArgumentParser(
        description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter
    )
    ap.add_argument("--codebook", action="append", type=Path, required=True,
                    help="data dictionary / codebook / CSV header (repeatable)")
    ap.add_argument("--context", action="append", default=[],
                    help="domain context: a review/guideline file (.md/.txt/.pdf) or a URL (repeatable)")
    ap.add_argument("--cohort-name", help="what the cohort is called")
    ap.add_argument("--out-dir", type=Path, default=Path("."))
    a = ap.parse_args()

    profile, context_text = build(a.codebook, a.context, a.cohort_name)
    if not profile["n_variables"]:
        raise SystemExit(
            "no variables found. If this is a data export, the first row must be the header; "
            "if it is a codebook, one column must be named variable / var / name / field / column."
        )

    a.out_dir.mkdir(parents=True, exist_ok=True)
    (a.out_dir / "cohort_profile.json").write_text(json.dumps(profile, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    (a.out_dir / "cohort_profile.md").write_text(render_markdown(profile), encoding="utf-8")
    if context_text:
        (a.out_dir / "context_extract.md").write_text(context_text + "\n", encoding="utf-8")

    print(f"{profile['n_variables']} variables enumerated from {len(a.codebook)} codebook(s)")
    for cluster, vs in profile["clusters"].items():
        print(f"  {cluster:<18} {len(vs)}")
    print(f"  serial groups      {len(profile['serial_groups'])}")
    print(f"  endpoint candidates {len(profile['endpoint_candidates'])}")
    unclassified = len(profile["clusters"].get("unclassified", []))
    if unclassified:
        print(f"\n{unclassified} variable(s) matched no cluster — review them; the lexicon is not exhaustive.")
    print(f"\nwrote {a.out_dir / 'cohort_profile.md'} + cohort_profile.json")
    print("ASK THE USER for: " + ", ".join(k.replace("_", " ") for k in profile["must_ask_user"]))
    return 0


if __name__ == "__main__":
    sys.exit(main())
